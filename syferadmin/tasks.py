from __future__ import absolute_import
import codecs
import csv
from datetime import datetime
import io
import os
import uuid

from django.conf import settings
from django.core.mail import EmailMultiAlternatives

from syferadmin.csv import UnicodeWriter
from syferadmin import video
from .signals import email_sent
from .utils import disconnect


try:
	from celery import shared_task
except ImportError:
	from syferadmin.celery import shared_task


@shared_task
def send_email(email, **kwargs):
	"""Send an email object"""
	if settings.Site.mode == 'prod':
		email.bcc += settings.SYFERADMIN_EMAIL_BCC
	if hasattr(send_email, 'request') and not getattr(send_email.request, 'called_directly', None) and not getattr(send_email.request, 'is_eager', None):
		disconnect()
	kwargs['email'] = email
	email_sent.send(sender=send_email, **kwargs)
	return email.send()


@shared_task
def encode_video(video_file, transfer_dir, formats):
	"Encode video into various formats using the encoder"
	video.encode(video_file, transfer_dir, formats)


@shared_task
def run_report(request_data, token, direct=False):
	run_start = datetime.now()
	from .reports import Report
	report = Report.get_report(request_data, token)
	detail = request_data.get('report').get('detail', '0') == '1'
	data = report.detail() if detail else report.summary()

	# If detail report run has taken longer than  60 seconds, the website stopped expecting a response
	# we need to save the report detail in a csv export, and mail it to the user
	if not direct:
		run_time = (datetime.now() - run_start).total_seconds()
		if detail and run_time > 60:
			export_result = do_export(report, data, token, 'csv')
			email_report_export(request_data['user'].email, report, export_result['filename'], request_data['region'], 'csv')

	return data


@shared_task
def export_report(request_data, token, export_as):
	run_start = datetime.now()
	from .reports import Report
	report = Report.get_report(request_data, token)

	data = report.detail()

	export_result = do_export(report, data, token, export_as)

	# If report run has taken longer than  60 seconds, the website stopped expecting a response
	# we need to mail it to the user instead
	run_time = (datetime.now() - run_start).total_seconds()
	if run_time > 60:
		email_report_export(request_data['user'].email, report, export_result['filename'], request_data['region'], export_as)

	return export_result


def do_export(report, data, token, export_as):
	directory = os.path.join(settings.MEDIA_ROOT, 'export')
	if not os.path.exists(directory):
		os.makedirs(directory)
	filename = ''
	if export_as == 'csv':
		filename = os.path.join(directory, '{}_{}.csv'.format(token, uuid.uuid1()))
		with codecs.open(filename, "w", 'utf-8') as fp:
			writer = UnicodeWriter(fp)

			fields = [i for i in report.fields if i[2] != 'image' and i[2] != 'html']

			writer.writerow([i[0].encode('utf-8') for i in fields])

			for record in data['results']:
				row = []
				for field in [i[1] for i in fields]:
					row.append(record[field])
				writer.writerow(row)
	else:
		import openpyxl
		from openpyxl.cell import get_column_letter
		from openpyxl.utils.units import pixels_to_points

		wb = openpyxl.Workbook()
		ws = wb.get_active_sheet()
		ws.title = token

		row_num = 1

		fields = [i for i in report.fields if i[2] != 'html']
		for col_idx, val in enumerate(fields):
			c = ws.cell(row=row_num, column=col_idx + 1)
			c.value = val[0]
			c.style.font.bold = True

		for record in data['results']:
			row_num += 1
			for col_idx, field in enumerate([i[1] for i in fields]):
				col_num = col_idx + 1
				if report.fields[col_idx][2] == 'image':
					try:
						image_file = record[field]
						if image_file.startswith(settings.MEDIA_URL):
							image_file = image_file[len(settings.MEDIA_URL) - 1:]
						if image_file:
							img = openpyxl.drawing.Image(settings.MEDIA_ROOT + image_file)
							img.anchor(ws.cell(row=row_num, column=col_num))
							ws.add_image(img)
							img_width = (img.drawing.width - 5) / 7
							col_width = ws.column_dimensions[get_column_letter(col_num)].width
							if not col_width or img_width > col_width:
								ws.column_dimensions[get_column_letter(col_num)].width = img_width
							row_height = ws.row_dimensions[row_num].height
							img_height = pixels_to_points(img.drawing.height + 10)  # img was overlapping. give it some padding
							if not row_height or img_height > row_height:
								ws.row_dimensions[row_num].height = img_height
					except(FileNotFoundError):
						pass
				else:
					c = ws.cell(row=row_num, column=col_num)
					c.value = record[field]
					c.style.alignment.wrap_text = True

		filename = os.path.join(directory, '{}_{}.xlsx'.format(token, uuid.uuid1()))
		wb.save(filename)

	return {'filename': filename, 'file_url': filename.replace(settings.MEDIA_ROOT.rstrip('/'), settings.MEDIA_URL.rstrip('/'))}


def email_report_export(email_address, report, filename, region, export_as):
	from .settings import Settings
	settings_ = Settings.load('values', region)
	email = EmailMultiAlternatives(report.title + ' Report export', 'The {} report you requested is attached'.format(report.title), '{} <{}>'.format(settings_.Company.name, settings_.Company.email), [email_address])
	if export_as == 'csv':
		# thanks to a bug in django that was fixed after the version we're currently using, attach_file can't attach a UTF-8 encoded csv file
		with codecs.open(filename, 'r', 'utf-8') as f:
			reader = csv.reader(f)
			csvfile = io.StringIO()
			csvwriter = UnicodeWriter(csvfile)
			for row in reader:
				csvwriter.writerow([i.encode('utf-8') for i in row])
			email.attach('{}.csv'.format(report.token), csvfile.getvalue(), 'application/csv')
	else:
		email.attach_file(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	send_email(email)
